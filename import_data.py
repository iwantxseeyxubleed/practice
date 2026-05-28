import shutil
from config import IMPORT_DIR, IMAGES_DIR
from database import get_or_create, table_empty

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def import_from_excel(con):
    if load_workbook is None:
        raise RuntimeError("Не установлен openpyxl. Выполните: pip install -r requirements.txt")

    IMAGES_DIR.mkdir(exist_ok=True)

    users_file = IMPORT_DIR / "user_import.xlsx"
    products_file = IMPORT_DIR / "Tovar.xlsx"
    points_file = IMPORT_DIR / "Пункты выдачи_import.xlsx"
    orders_file = IMPORT_DIR / "Заказ_import.xlsx"

    if table_empty(con, "users") and users_file.exists():
        ws = load_workbook(users_file, data_only=True).active
        for role, full_name, login, password in ws.iter_rows(min_row=2, values_only=True):
            if not full_name:
                continue
            role_id = get_or_create(con, "roles", str(role).strip())
            con.execute(
                "INSERT OR IGNORE INTO users(role_id, full_name, login, password) VALUES(?,?,?,?)",
                (role_id, str(full_name).strip(), str(login).strip(), str(password).strip()),
            )

    if table_empty(con, "pickup_points") and points_file.exists():
        ws = load_workbook(points_file, data_only=True).active
        for (address,) in ws.iter_rows(values_only=True):
            if address:
                con.execute("INSERT OR IGNORE INTO pickup_points(address) VALUES(?)", (str(address).strip(),))

    if table_empty(con, "products") and products_file.exists():
        ws = load_workbook(products_file, data_only=True).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            article, name, unit, price, supplier, manufacturer, category, discount, stock, description, photo = row
            if not article:
                continue
            supplier_id = get_or_create(con, "suppliers", str(supplier).strip())
            manufacturer_id = get_or_create(con, "manufacturers", str(manufacturer).strip())
            category_id = get_or_create(con, "categories", str(category).strip())
            photo_name = str(photo).strip() if photo else "picture.png"
            source = IMPORT_DIR / photo_name
            target = IMAGES_DIR / photo_name
            if source.exists() and not target.exists():
                shutil.copy(source, target)
            con.execute(
                """
                INSERT OR IGNORE INTO products(article, name, unit, price, supplier_id, manufacturer_id,
                category_id, discount, stock_quantity, description, image_path)
                VALUES(?,?,?,?,?,?,?,?,?,?,?)
                """,
                (str(article).strip(), str(name).strip(), str(unit).strip(), float(price or 0), supplier_id,
                 manufacturer_id, category_id, int(discount or 0), int(stock or 0), str(description or ""), str(target)),
            )

    if table_empty(con, "orders") and orders_file.exists():
        ws = load_workbook(orders_file, data_only=True).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            order_id, item_text, order_date, delivery_date, point_id, client, code, status = row[:8]
            if not order_id:
                continue
            con.execute(
                """
                INSERT OR IGNORE INTO orders(id, order_date, delivery_date, pickup_point_id, client_full_name, receive_code, status)
                VALUES(?,?,?,?,?,?,?)
                """,
                (int(order_id), str(order_date.date()) if hasattr(order_date, "date") else str(order_date),
                 str(delivery_date.date()) if hasattr(delivery_date, "date") else str(delivery_date),
                 int(point_id) if point_id else None, str(client or ""), int(code or 0), str(status or "")),
            )
            parts = [p.strip() for p in str(item_text or "").split(",")]
            for i in range(0, len(parts) - 1, 2):
                article = parts[i]
                try:
                    qty = int(parts[i + 1])
                except ValueError:
                    qty = 1
                exists = con.execute("SELECT 1 FROM products WHERE article=?", (article,)).fetchone()
                if exists:
                    con.execute(
                        "INSERT INTO order_items(order_id, product_article, quantity) VALUES(?,?,?)",
                        (int(order_id), article, qty),
                    )
    con.commit()
